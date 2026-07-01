from pathlib import Path
import openpyxl
import pytest


@pytest.fixture
def evidence_root(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws["A1"] = "user_id"; ws["B1"] = "name"; ws["C1"] = "role"
    ws["A2"] = "14"; ws["B2"] = "Sam Lee"; ws["C2"] = "payments_admin"
    wb.save(tmp_path / "access_export.xlsx")
    (tmp_path / "approvals.txt").write_text(
        "Approval 1042: user 14 approved access read_only\n", encoding="utf-8"
    )
    return tmp_path
